#!/usr/bin/env python3
"""Extract sheet contents from an Excel (.xlsx) file as TSV."""

import sys

from openpyxl import load_workbook


def main(path: str) -> int:
    workbook = load_workbook(path, data_only=True, read_only=True)
    for sheet in workbook.worksheets:
        print(f'--- sheet "{sheet.title}" ---')
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if value is None else str(value) for value in row]
            while cells and cells[-1] == "":
                cells.pop()
            if cells:
                print("\t".join(cells))
        print()
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: extract.py <file.xlsx>", file=sys.stderr)
        sys.exit(2)
    sys.exit(main(sys.argv[1]))
